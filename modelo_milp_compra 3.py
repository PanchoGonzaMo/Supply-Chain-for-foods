# -*- coding: utf-8 -*-
"""
================================================================================
 modelo_milp_compra.py
 Modelo de optimización de compra de productos importados — IKEA Food Chile
 Tesis Magíster en Dirección de Operaciones y Logística (MOL · UNAB)
 Francisco González — "Diseño de un Modelo de Compra de Productos Importados
 en IKEA Food Chile"
================================================================================

Este módulo implementa el MODELO MILP descrito en el Capítulo 4.4 y operado en
el Capítulo 9 de la tesis. La formulación es una adaptación de Pauls-Worm et al.
(2014) con las tres extensiones propias del caso IKEA Food: (i) pedido en pallets
completos, (ii) disyunción de capacidad de contenedor (23 o 40–44 pallets) y
(iii) vida útil residual al momento del arribo.

El MILP consume los INSUMOS generados por los modelos intermedios (Capítulo 6):
    - D̂_{p,t}  : forecast de demanda. Se corrige por el SESGO sistemático de
                 +23 % detectado al lag operativo de 6 meses (sección 8).
    - V_{p,t}  : vencimientos proyectados del stock vigente (modelo de
                 obsolescencia + FEFO retroactivo, Ec. 13).
    - I_{p,0}  : inventario consolidado inicial (CD + tiendas).
    - A_{p,t}  : arribos ya confirmados en tránsito (proyección de inventario).
    - σ_p      : desviación estándar del error de pronóstico = RMSE (sección 8),
                 que parametriza el stock de seguridad SS = Z · σ · √L (Ec. 4).

------------------------------- FORMULACIÓN ------------------------------------
Conjuntos
    p ∈ P   productos (SKU importados)
    t ∈ {1..T}   períodos (meses); horizonte T = 12

Variables de decisión
    n_{p,t} ∈ ℤ₊      nº de pallets pedidos                         (Ec. 9)
    Q_{p,t} = n·CAP_p unidades pedidas                              (Ec. 9)
    I_{p,t} ≥ 0       inventario al cierre del período t
    S_{p,t} ≥ 0       quiebre (demanda no satisfecha, venta perdida)
    δ_{p,t} ∈ {0,1}   1 si se genera un pedido en t
    k^s_{p,t} ∈ ℤ₊    nº de contenedores CHICOS (23 pallets exactos c/u)  (Ec. 10)
    k^f_{p,t} ∈ ℤ₊    nº de contenedores LLENOS (40–44 pallets c/u)       (Ec. 10)
    F_{p,t} ∈ ℤ₊      pallets embarcados en los contenedores llenos       (Ec. 10)

Función objetivo                                                    (Ec. 7)
    min  Σ_{p,t} [ h_p·I_{p,t} + p_p·S_{p,t} + o·δ_{p,t} + c_u,p·Q_{p,t} ]
    (el término w_p·V_{p,t} es constante porque V es dato del modelo de
     obsolescencia; se reporta pero no altera el óptimo. Bajo CIP el flete NO
     entra en la función objetivo.)

Sujeto a
    Balance con vencimientos y quiebre (extensión de Ec. 8):
        I_{p,t} = I_{p,t-1} + A_{p,t} + Q_{p,t-L} − D_{p,t} + S_{p,t} − V_{p,t}
    Stock de seguridad (Ec. 5, σ = RMSE, Ec. 4):
        I_{p,t} ≥ SS_{p,k}            (soft por defecto para robustez)
    Integralidad de pallet (Ec. 9):
        Q_{p,t} = n_{p,t} · CAP_p
    Capacidad de contenedor (Ec. 10) — el pedido es una COMBINACIÓN de
    contenedores, no una cantidad libre. El chico transporta EXACTAMENTE 23
    pallets (ni más ni menos); el lleno, entre 40 y 44:
        n_{p,t} = 23·k^s_{p,t} + F_{p,t}
        40·k^f_{p,t} ≤ F_{p,t} ≤ 44·k^f_{p,t}
    Cantidades alcanzables: 0, 23, 40–44, 46, 63–67, 69, 80–90, 92, ...
    Huecos NO alcanzables:  1–22, 24–39, 45, 47–62, 68, ...
    Esta formulación habilita pedidos > 44 pallets (varios contenedores), caso
    que la disyunción original truncaba silenciosamente en 44.
    Vida útil residual al arribo (Ec. 11):
        I_{p,t-1} + A_{p,t} + Q_{p,t-L} ≤ Σ_{k=0}^{VU_p-1} D̂_{p,t+k}

Solver: PuLP + CBC (código abierto), tal como se especifica en la sección 5.3.
================================================================================
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import pulp


# =============================================================================
# 1. PARÁMETROS QUE PROVIENEN DE LOS MODELOS INTERMEDIOS Y DEL ANÁLISIS (Cap. 8)
# =============================================================================

# Nivel de servicio diferenciado por clase ABC → factor Z de la normal estándar
#   A: 98 %  ·  B: 95 %  ·  C: 90 %   (Tabla 3 de la tesis)
Z_BY_CLASS: Dict[str, float] = {"A": 2.054, "B": 1.645, "C": 1.282}

# Sesgo sistemático del forecast nacional al lag operativo de 6 meses (sección 8).
#   El forecast sobreestima la demanda en +23,2 %  →  se corrige antes de optimizar.
DEFAULT_FORECAST_BIAS: float = 0.232

# Lead time completo de importación (orden → liberación SEREMI), en meses.
DEFAULT_LEAD_TIME: int = 6


def safety_stock(z: float, sigma: float, lead_time: int) -> float:
    """Stock de seguridad — Ec. (4):  SS = Z · σ · √L.

    σ es la desviación estándar del error de pronóstico, estimada por el RMSE
    del análisis de precisión por lag (sección 8). L es el lead time en períodos.
    """
    return z * sigma * math.sqrt(lead_time)


def debias_forecast(forecast: np.ndarray, bias: float = DEFAULT_FORECAST_BIAS) -> np.ndarray:
    """Corrige el sesgo sistemático del forecast nacional.

    Si el pronóstico sobreestima en una fracción `bias` (p. ej. +0,232),
    la demanda esperada insesgada es  D̂_insesgada = D̂_forecast / (1 + bias).
    Esta corrección es el mecanismo por el cual el hallazgo de la sección 8
    (sobre-pronóstico → sobrestock → vencimientos) se traduce en menos compra.
    """
    return np.asarray(forecast, dtype=float) / (1.0 + bias)


def extend_demand(
    forecast_12: np.ndarray,
    historico_mensual: Optional[pd.Series] = None,
    meses_extra: int = 6,
) -> np.ndarray:
    """Extiende la demanda más allá del forecast (COLA TÉCNICA, no un pronóstico).

    ¿Por qué hace falta? La restricción de vida útil (Ec. 11) mira HACIA ADELANTE:
    para autorizar un pedido que arriba en el mes 12 necesita saber qué demanda hay
    en los meses 12..12+VU. Si el forecast termina en el mes 12, el modelo "ve" cero
    demanda futura y BLOQUEA los pedidos tardíos —un artefacto, no una decisión
    económica—.

    Método: seasonal naive. Se repite el mismo mes del ciclo anual anterior tomándolo
    del forecast disponible. Si hay consumo histórico, se usa su perfil estacional.
    Se documenta como supuesto: la cola NO es una proyección de negocio.
    """
    f = np.asarray(forecast_12, dtype=float)
    n = len(f)
    if meses_extra <= 0:
        return f
    if historico_mensual is not None and len(historico_mensual) >= 12:
        # perfil estacional del histórico, escalado al nivel del forecast
        h = np.asarray(historico_mensual, dtype=float)[-12:]
        perfil = h / h.mean() if h.mean() > 0 else np.ones(12)
        nivel = f.mean()
        cola = np.array([nivel * perfil[(n + k) % 12] for k in range(meses_extra)])
    else:
        # seasonal naive puro: repite el ciclo del propio forecast
        cola = np.array([f[(n + k) % n] for k in range(meses_extra)])
    return np.concatenate([f, cola])


# =============================================================================
# 2. ESTRUCTURAS DE DATOS
# =============================================================================

@dataclass
class SKU:
    """Parámetros de un producto importado (una fila de la tabla del Bloque 3)."""
    sku_id: str
    abc_class: str                  # "A" | "B" | "C"
    units_per_pallet: int           # CAP_p  (Ec. 9) — unidades base por pallet
    units_per_case: int = 1         # unidades base por CAJA (para el output a Compras)
    shelf_life_months: int = 12     # VU_p   (Ec. 11) — vida útil residual al arribo
    holding_cost: float = 0.0       # h_p   ($/unidad/período) = tarifa bodega / CAP_p
    unit_cost: float = 0.0          # c_u,p ($/unidad)  (Dynamics)
    stockout_penalty: float = 0.0   # p_p   ($/unidad)  ≈ precio de venta
    expiry_cost: float = 0.0        # w_p   ($/unidad)  costo medio al vencimiento
    rmse: float = 0.0               # σ_p   (= RMSE del forecast, sección 8)
    init_inventory: float = 0.0     # I_{p,0}
    # ---- Vida útil residual al arribo (Ec. 11), en meses ---------------------
    # Se toma del REGISTRO DE IMPORTACIONES (vida útil restante medida al arribo),
    # no de la vida útil nominal de fábrica. Se manejan dos escenarios:
    #   · shelf_life_w   : promedio PONDERADO por unidades arribadas  → escenario BASE
    #   · shelf_life_p25 : percentil conservador (P25)                → escenario CONSERVADOR
    # `shelf_life_months` es el valor efectivamente usado en la corrida en curso.
    shelf_life_w: Optional[float] = None
    shelf_life_p25: Optional[float] = None
    # Grupo de consolidación de contenedor: (origen, tipo de producto).
    # Los pallets de todos los SKU de un mismo grupo comparten contenedor; la
    # estiba (23 o 40-44) se cumple a nivel de EMBARQUE, no de SKU individual.
    # Proveedor único (IKEA Suecia) ⇒ el criterio real es el TIPO (seco/congelado).
    container_group: str = "SECO"   # p.ej. "SECO" | "CONGELADO"

    def z(self) -> float:
        return Z_BY_CLASS[self.abc_class]


@dataclass
class ModelConfig:
    """Configuración global del MILP."""
    horizon: int = 12               # T (meses)
    lead_time: int = DEFAULT_LEAD_TIME   # L (meses)
    forecast_bias: float = DEFAULT_FORECAST_BIAS
    order_fixed_cost: float = 50_000.0   # o  ($/pedido)
    # --- Costo de bodega: el operador cobra por PALLET-DÍA -------------------
    # Cuenta los pallets almacenados cada día, suma los pallet-día del corte y
    # los multiplica por la tarifa diaria en UF. Por lo tanto el costo NO es
    # proporcional a las unidades, sino a las POSICIONES DE PALLET ocupadas:
    # un pallet con 50 unidades cuesta lo mismo que uno con 1.200.
    tarifa_uf_pallet_dia: float = 0.3019   # UF por pallet por día
    uf_clp: float = 40_844.79              # valor UF (cambia a diario)
    dias_mes: float = 30.44
    cobro_por_pallet: bool = True          # True = costo por posición (redondeo ↑)

    def tarifa_pallet_mes(self) -> float:
        """$ por posición de pallet por mes."""
        return self.tarifa_uf_pallet_dia * self.dias_mes * self.uf_clp
    # Patrones de ESTIBA permitidos por el proveedor (Ec. 10).
    # Es SIEMPRE el mismo contenedor grande: lo que cambia es cómo se estiba la
    # carga por temas de balance. Un contenedor va con 23 pallets exactos, o
    # bien con 40 a 44 pallets. No existen configuraciones intermedias.
    stow_23: int = 23                # estiba de 23 pallets (exactos)
    stow_full_min: int = 40          # estiba completa: mínimo
    stow_full_max: int = 44          # estiba completa: máximo
    # Sin tope de contenedores: el proveedor acepta los que se pidan.
    # (cota superior técnica para el solver, no una restricción de negocio)
    max_containers_each: int = 20
    big_m_pallets: int = 500         # cota técnica para la bandera de pedido
    # Stock de seguridad como restricción dura o blanda (penalizada)
    # --- OBJETIVO -----------------------------------------------------------
    # "stock" : minimiza pallet-mes de inventario sujeto a nivel de servicio.
    #           Como la tarifa de bodega es la misma para toda posición, minimizar
    #           pallet-mes EQUIVALE a minimizar el gasto de bodega (la tarifa es
    #           una constante multiplicativa). El nivel de servicio deja de ser un
    #           costo y pasa a ser RESTRICCIÓN: no hay que calibrar penalizaciones.
    # "costo" : función de costo clásica (Ec. 7). Requiere h, p, o, c_u.
    objetivo: str = "stock"
    # Meses de demanda estimada MÁS ALLÁ del forecast (cola técnica para la Ec. 11).
    # Marca cuántos períodos finales del horizonte NO provienen del forecast real.
    meses_cola: int = 6
    # Horizonte rodante: solo el pedido del PRIMER período se compromete; el resto es
    # indicativo y se recalcula el mes siguiente con forecast actualizado.
    solo_primer_pedido_compromete: bool = True
    hard_safety_stock: bool = False
    ss_shortfall_penalty: float = 1.0e3   # penalización por unidad bajo SS (modo blando)
    stockout_big_m: float = 1.0e4         # el quiebre es inaceptable: se evita siempre
                                          # que sea factible (nivel de servicio = restricción)
    solver_msg: bool = False
    # Límite de tiempo del solver (s) y brecha de optimalidad aceptable.
    # En MILP grandes conviene aceptar una solución con brecha pequeña antes que
    # esperar el óptimo exacto: la diferencia práctica es despreciable.
    time_limit: int = 120
    gap_rel: float = 0.01          # 1 % de brecha


# =============================================================================
# 3. CONSTRUCCIÓN Y RESOLUCIÓN DEL MILP
# =============================================================================

def solve_milp(
    skus: List[SKU],
    forecast: Dict[str, np.ndarray],          # D̂_{p,t} crudo (con sesgo), shape [T]
    expirations: Dict[str, np.ndarray],       # V_{p,t}  (modelo de obsolescencia)
    arrivals: Optional[Dict[str, np.ndarray]] = None,   # A_{p,t} arribos confirmados
    cfg: ModelConfig = ModelConfig(),
    disponibilidad: Optional[Dict[str, np.ndarray]] = None,  # 1/0: proveedor puede despachar
) -> Dict:
    """Construye y resuelve el modelo MILP de compra.

    Retorna un diccionario con el estado, el costo óptimo y un DataFrame de
    resultados por SKU y período (la tabla que la tesis escribe de vuelta a
    BigQuery para el área de compras).
    """
    T, L = cfg.horizon, cfg.lead_time
    periods = range(1, T + 1)

    prob = pulp.LpProblem("Compra_IKEA_Food_MILP", pulp.LpMinimize)

    # grupos de consolidación de contenedor (origen + tipo de producto)
    groups = sorted({sk.container_group for sk in skus})
    by_group = {g: [sk.sku_id for sk in skus if sk.container_group == g] for g in groups}

    # ---- variables ----
    n, Q, I, S, delta, k_s, k_f, F, ss_short, Pst = ({} for _ in range(10))
    # contenedores: se deciden por GRUPO y período (un contenedor lleva varios SKU)
    for g in groups:
        for t in periods:
            k_s[g, t] = pulp.LpVariable(f"ks_{g}_{t}", lowBound=0,
                                        upBound=cfg.max_containers_each, cat="Integer")
            k_f[g, t] = pulp.LpVariable(f"kf_{g}_{t}", lowBound=0,
                                        upBound=cfg.max_containers_each, cat="Integer")
            F[g, t]   = pulp.LpVariable(f"F_{g}_{t}", lowBound=0, cat="Integer")
    for sk in skus:
        p = sk.sku_id
        for t in periods:
            n[p, t]     = pulp.LpVariable(f"n_{p}_{t}", lowBound=0, cat="Integer")
            Q[p, t]     = pulp.LpVariable(f"Q_{p}_{t}", lowBound=0, cat="Continuous")
            I[p, t]     = pulp.LpVariable(f"I_{p}_{t}", lowBound=0, cat="Continuous")
            S[p, t]     = pulp.LpVariable(f"S_{p}_{t}", lowBound=0, cat="Continuous")
            delta[p, t] = pulp.LpVariable(f"d_{p}_{t}", cat="Binary")
            ss_short[p, t] = pulp.LpVariable(f"sss_{p}_{t}", lowBound=0, cat="Continuous")
            # posiciones de pallet ocupadas en bodega (entero: se cobra el pallet
            # completo aunque esté parcialmente lleno)
            Pst[p, t]      = pulp.LpVariable(f"P_{p}_{t}", lowBound=0, cat="Integer")

    # ---- demanda insesgada y stock de seguridad por SKU ----
    Dhat = {sk.sku_id: debias_forecast(forecast[sk.sku_id], cfg.forecast_bias) for sk in skus}
    SS   = {sk.sku_id: safety_stock(sk.z(), sk.rmse, L) for sk in skus}

    # ---- función objetivo ----
    obj = []
    for sk in skus:
        p = sk.sku_id
        for t in periods:
            if cfg.objetivo == "stock":
                # MINIMIZAR INVENTARIO en pallet-mes (normaliza entre SKU con
                # distinta unidad base: KG vs PC no son comparables; pallets sí).
                # El gasto de bodega cae orgánicamente al bajar el stock.
                obj.append(Pst[p, t])
                # el quiebre NO es un costo: es una violación del nivel de servicio
                obj.append(cfg.stockout_big_m * S[p, t])
            else:
                # Modo costo: solo bodega + servicio. NO incluye costo fijo de pedido
                # (o·δ) ni costo unitario de compra (c_u·Q): la compra total en el
                # horizonte está fijada por la demanda, de modo que c_u·Q es una
                # constante y o·δ no corresponde al objetivo de esta tesis.
                obj.append(cfg.tarifa_pallet_mes() * Pst[p, t]
                           if cfg.cobro_por_pallet else sk.holding_cost * I[p, t])
                obj.append(sk.stockout_penalty * S[p, t])
            if not cfg.hard_safety_stock:
                obj.append(cfg.ss_shortfall_penalty * ss_short[p, t])
    prob += pulp.lpSum(obj)

    # ---- restricciones ----
    for sk in skus:
        p = sk.sku_id
        CAP, VU = sk.units_per_pallet, sk.shelf_life_months
        A = (arrivals or {}).get(p, np.zeros(T))
        for t in periods:
            inv_prev = sk.init_inventory if t == 1 else I[p, t - 1]
            arr_order = Q[p, t - L] if (t - L) >= 1 else 0   # pedido que arriba en t
            arr_conf  = float(A[t - 1])                       # arribo ya confirmado

            # Balance con vencimientos y quiebre (extensión de Ec. 8)
            prob += (
                I[p, t] == inv_prev + arr_conf + arr_order
                          - Dhat[p][t - 1] + S[p, t] - float(expirations[p][t - 1]),
                f"balance_{p}_{t}",
            )

            # Stock de seguridad (Ec. 5);  σ = RMSE (Ec. 4)
            if cfg.hard_safety_stock:
                prob += I[p, t] >= SS[p], f"ss_{p}_{t}"
            else:
                prob += I[p, t] + ss_short[p, t] >= SS[p], f"ss_{p}_{t}"

            # Integralidad de pallet (Ec. 9)
            prob += Q[p, t] == CAP * n[p, t], f"pallet_{p}_{t}"

            # Posiciones de pallet ocupadas en bodega (Ec. 12 — nueva).
            # P ≥ I/CAP con P entero ⇒ P = ⌈I/CAP⌉ en el óptimo (el modelo nunca
            # paga de más). Refleja que el operador cobra la posición completa.
            prob += CAP * Pst[p, t] >= I[p, t], f"palletpos_{p}_{t}"

            # bandera de pedido del SKU: δ = 1 si el SKU pide pallets en t
            prob += n[p, t] <= cfg.big_m_pallets * delta[p, t], f"order_ub_{p}_{t}"

            # Disponibilidad en origen: no se puede pedir lo que el proveedor no tiene
            if disponibilidad is not None and p in disponibilidad:
                if float(disponibilidad[p][t - 1]) <= 0:
                    prob += n[p, t] == 0, f"nodisp_{p}_{t}"

            # No tiene sentido pedir si el arribo cae FUERA del horizonte
            if t + L > T:
                prob += n[p, t] == 0, f"fuera_horiz_{p}_{t}"
            prob += n[p, t] >= delta[p, t], f"order_lb_{p}_{t}"

            # Vida útil residual al arribo (Ec. 11): el stock disponible tras el
            # arribo no puede superar la demanda consumible dentro de la vida útil.
            window = [Dhat[p][k - 1] for k in range(t, min(t + VU, T + 1))]
            prob += inv_prev + arr_conf + arr_order <= pulp.lpSum(window), f"shelf_{p}_{t}"

    # ---- Estiba de contenedor (Ec. 10) a nivel de EMBARQUE ----
    # Un contenedor transporta pallets de VARIOS SKU del mismo grupo (origen +
    # tipo de producto). El proveedor solo admite dos patrones de estiba por
    # balance de carga: 23 pallets exactos, o de 40 a 44 pallets por contenedor.
    # Se admiten tantos contenedores como haga falta (sin tope).
    #     Σ_p n_{p,t} = 23·k_s + F      con   40·k_f ≤ F ≤ 44·k_f
    for g in groups:
        for t in periods:
            pallets_grupo = pulp.lpSum(n[p, t] for p in by_group[g])
            prob += pallets_grupo == cfg.stow_23 * k_s[g, t] + F[g, t], f"cont_mix_{g}_{t}"
            prob += F[g, t] >= cfg.stow_full_min * k_f[g, t], f"cont_flb_{g}_{t}"
            prob += F[g, t] <= cfg.stow_full_max * k_f[g, t], f"cont_fub_{g}_{t}"

    # ---- resolver con CBC ----
    prob.solve(pulp.PULP_CBC_CMD(msg=cfg.solver_msg,
                                timeLimit=cfg.time_limit,
                                gapRel=cfg.gap_rel))
    status = pulp.LpStatus[prob.status]

    # ---- extraer resultados ----
    rows = []
    for sk in skus:
        p = sk.sku_id
        for t in periods:
            npal = int(round(n[p, t].value() or 0))
            qty  = (Q[p, t].value() or 0.0)
            g    = sk.container_group
            ns   = int(round(k_s[g, t].value() or 0))   # contenedores estibados a 23 (del grupo)
            nf   = int(round(k_f[g, t].value() or 0))   # contenedores estibados a 40-44 (del grupo)
            partes = ([f"{ns}×estiba 23"] if ns else []) + ([f"{nf}×estiba 40-44"] if nf else [])
            cont = " + ".join(partes) if partes else "—"
            rows.append({
                "sku": p, "clase": sk.abc_class, "grupo": sk.container_group, "periodo": t,
                "pallets": npal, "unidades_pedidas": round(qty),
                "contenedor_grupo": cont, "cont_estiba23": ns, "cont_estiba40_44": nf,
                "n_contenedores": ns + nf,
                "inventario_fin": round(I[p, t].value() or 0.0),
                "pallets_bodega": int(round(Pst[p, t].value() or 0)),
                "costo_bodega": round((Pst[p, t].value() or 0) * cfg.tarifa_pallet_mes()),
                "stock_seguridad": round(SS[p]),
                "quiebre_u": round(S[p, t].value() or 0.0),
                "vencimiento_u": round(float(expirations[p][t - 1])),
                "demanda_insesgada": round(Dhat[p][t - 1]),
            })
    df = pd.DataFrame(rows)

    return {
        "status": status,
        "costo_total": pulp.value(prob.objective),
        "resultados": df,
        "safety_stock": SS,
        "demanda_insesgada": Dhat,
        "config": cfg,
        "skus": {s.sku_id: s for s in skus},
    }


# =============================================================================
# 4. VISUALIZACIÓN DE RESULTADOS (gráfico del modelo por SKU)
# =============================================================================

def plot_sku(result: Dict, sku_id: str, path: Optional[str] = None):
    """Grafica la trayectoria de inventario, el piso de stock de seguridad,
    los pedidos (arribos) y los vencimientos para un SKU."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import font_manager
    for fp in ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
               "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"]:
        try: font_manager.fontManager.addfont(fp)
        except Exception: pass
    plt.rcParams["font.family"] = "Liberation Sans"
    plt.rcParams["axes.unicode_minus"] = False
    NAVY, IKEA, YELL, RED = "#1F3864", "#0058A3", "#FFCC00", "#C0392B"

    df = result["resultados"]
    d = df[df["sku"] == sku_id].sort_values("periodo")
    cfg = result["config"]; L = cfg.lead_time
    t = d["periodo"].values

    fig, ax1 = plt.subplots(figsize=(9, 4.6))
    ax1.plot(t, d["inventario_fin"], "-o", color=IKEA, lw=2.2, ms=5,
             mfc="white", mec=IKEA, mew=1.5, label="Inventario proyectado", zorder=4)
    ax1.axhline(d["stock_seguridad"].iloc[0], color=RED, ls="--", lw=1.4,
                label=f"Stock de seguridad (Z·σ·√L = {d['stock_seguridad'].iloc[0]:,.0f})")

    # arribos = pedidos desplazados por el lead time
    arr = np.zeros(len(t))
    for _, r in d.iterrows():
        ta = int(r["periodo"]) + L
        if 1 <= ta <= len(t):
            arr[ta - 1] += r["unidades_pedidas"]
    ax1.bar(t, arr, width=0.55, color=YELL, ec=NAVY, lw=0.6, alpha=0.85,
            label=f"Arribos (pedidos, L={L})", zorder=2)

    venc = d["vencimiento_u"].values
    if venc.sum() > 0:
        ax1.bar(t, -venc, width=0.55, color=RED, ec="#7B241C", lw=0.6, alpha=0.7,
                label="Vencimientos", zorder=2)

    ax1.axhline(0, color=NAVY, lw=0.8)
    ax1.set_xlabel("Período (mes del horizonte)", color=NAVY)
    ax1.set_ylabel("Unidades", color=NAVY)
    ax1.set_xticks(t)
    ax1.grid(axis="y", color="#E8E8E8", lw=0.8); ax1.set_axisbelow(True)
    for s in ["top", "right"]: ax1.spines[s].set_visible(False)
    cls = result["skus"][sku_id].abc_class
    ax1.set_title(f"Modelo MILP — SKU {sku_id} (clase {cls}): plan de compra y trayectoria de inventario",
                  color=NAVY, fontweight="bold", loc="left", pad=10)
    ax1.legend(loc="upper right", fontsize=8.5, frameon=False, ncol=2)
    fig.tight_layout()
    if path:
        fig.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    return fig


def pedido_del_mes(result: Dict) -> pd.DataFrame:
    """Extrae EL PEDIDO A EJECUTAR: solo el del período 1.

    Proceso real: cada mes el equipo comercial entrega un forecast a 12 meses y se
    calcula el pedido de ESE mes. El pedido llega L meses después y solo debe cubrir
    hasta que arribe el siguiente (≈ un mes de demanda + stock de seguridad), porque
    el mes próximo se vuelve a pedir.

    ⚠️ Por eso el MILP resuelve el horizonte COMPLETO aunque solo se ejecute el
    primer pedido: si al modelo se le prohibiera pedir en los meses siguientes,
    creería que nunca más puede reponer y COMPRARÍA PARA TODO EL HORIZONTE,
    llenando la bodega. Las compras futuras son variables de andamiaje —el supuesto
    de que "el mes que viene también se pedirá"—, no un plan a comprometer.
    """
    df, cfg = result["resultados"], result["config"]
    skus = result["skus"]
    Dhat = result["demanda_insesgada"]

    hoy = df[(df["periodo"] == 1) & (df["pallets"] > 0)].copy()
    if hoy.empty:
        return pd.DataFrame(columns=["sku", "pallets", "cajas", "unidades_base",
                                     "meses_de_venta", "contenedor", "grupo"])

    filas = []
    for _, r in hoy.iterrows():
        p = r["sku"]; sk = skus[p]
        u = float(r["unidades_pedidas"])
        dem_mes = float(np.mean(Dhat[p])) if np.mean(Dhat[p]) > 0 else 0.0
        filas.append({
            "sku": p,
            "descripcion": getattr(sk, "descripcion", ""),
            "clase_abc": sk.abc_class,
            "grupo": sk.container_group,
            "pallets": int(r["pallets"]),
            "cajas": round(u / sk.units_per_case, 1) if sk.units_per_case else 0,
            "unidades_base": int(u),
            "meses_de_venta": round(u / dem_mes, 2) if dem_mes > 0 else np.nan,
            "mes_arribo_estimado": 1 + cfg.lead_time,
            "contenedor": r["contenedor_grupo"],
        })
    out = pd.DataFrame(filas).sort_values(["grupo", "sku"])

    # resumen de contenedores del embarque (a nivel de grupo, no de SKU)
    print("PEDIDO A EJECUTAR ESTE MES")
    print(f"  arribo estimado: en {cfg.lead_time} meses (liberación SEREMI)")
    for g, gg in out.groupby("grupo"):
        pal = gg["pallets"].sum()
        cont = df[(df["periodo"] == 1) & (df["grupo"] == g)]["contenedor_grupo"].iloc[0]
        print(f"  {g:18s} {pal:>4} pallets → {cont}")
    return out


def export_excel(result: Dict, path: str = "plan_compra_milp.xlsx",
                 results_all: Optional[Dict[str, Dict]] = None) -> str:
    """Exporta el plan de compra a Excel para el área de Compras.

    El archivo es la HOJA DE TRABAJO del comprador: trae el pedido sugerido en
    cajas, pallets y unidad base, más los MESES DE VENTA que cubre, de modo que
    el pedido pueda ajustarse manualmente (por razones comerciales o quiebres del
    proveedor) ANTES de cargarlo en la plataforma del proveedor.

    Hojas:
      · Plan de compra  — una fila por SKU y período con pedido (editable)
      · Detalle mensual — trayectoria completa de inventario por SKU y mes
      · Parámetros      — supuestos del modelo (trazabilidad)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df, cfg = result["resultados"], result["config"]
    skus = result["skus"]
    Dhat = result["demanda_insesgada"]

    # demanda mensual promedio insesgada → base para "meses de venta"
    dem_mes = {s: float(np.mean(v)) if np.mean(v) > 0 else 0.0 for s, v in Dhat.items()}

    NAVY = "1F3864"; HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY = Font(name="Arial", size=10)
    FILL = PatternFill("solid", fgColor="0058A3")
    EDIT = PatternFill("solid", fgColor="FFF2CC")   # amarillo: celdas ajustables
    thin = Side(style="thin", color="BFBFBF"); BORD = Border(thin, thin, thin, thin)

    wb = Workbook(); ws = wb.active; ws.title = "Plan de compra"

    ws["A1"] = "PLAN DE COMPRA SUGERIDO — Modelo MILP · IKEA Food Chile"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=NAVY)
    ws["A2"] = (f"Lead time {cfg.lead_time} meses · horizonte {cfg.horizon} meses · "
                f"forecast corregido por sesgo +{cfg.forecast_bias:.1%} · "
                f"estiba permitida: 23 pallets o 40-44 pallets por contenedor")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A3"] = ("Las celdas en amarillo son AJUSTABLES: modifique el pedido por razones "
                "comerciales o quiebres del proveedor antes de cargarlo en la plataforma.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="C0392B")

    cols = ["SKU", "Clase ABC", "Grupo contenedor", "Mes de pedido", "Mes de arribo",
            "Pallets", "Cajas", "Unidades base",
            "Meses de venta que cubre", "Contenedores", "Estiba",
            "Inventario proyectado", "Stock de seguridad",
            "PEDIDO AJUSTADO (pallets)", "Comentario Compras"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=5, column=j, value=c)
        cell.font = HEAD; cell.fill = FILL; cell.border = BORD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ped = df[df["pallets"] > 0].sort_values(["sku", "periodo"])
    r = 6
    for _, row in ped.iterrows():
        s = row["sku"]; sk = skus[s]
        u = float(row["unidades_pedidas"])
        cajas = u / sk.units_per_case if sk.units_per_case else 0
        meses_venta = u / dem_mes[s] if dem_mes.get(s, 0) > 0 else 0
        vals = [s, row["clase"], row["grupo"], int(row["periodo"]), int(row["periodo"]) + cfg.lead_time,
                int(row["pallets"]), round(cajas, 1), int(u),
                round(meses_venta, 2), int(row["n_contenedores"]), row["contenedor_grupo"],
                int(row["inventario_fin"]), int(row["stock_seguridad"]),
                int(row["pallets"]), ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = BODY; cell.border = BORD
            if j in (14, 15):                      # columnas editables por Compras
                cell.fill = EDIT
            if j in (7, 9):
                cell.number_format = "#,##0.0"
            elif j in (6, 8, 12, 13, 14):
                cell.number_format = "#,##0"
        r += 1

    widths = [20, 10, 16, 13, 13, 9, 10, 14, 14, 13, 22, 15, 15, 16, 26]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A6"

    # ---- hoja 2: detalle mensual ----
    ws2 = wb.create_sheet("Detalle mensual")
    det = ["sku", "clase", "grupo", "periodo", "demanda_insesgada", "inventario_fin",
           "stock_seguridad", "pallets", "unidades_pedidas", "quiebre_u", "vencimiento_u"]
    for j, c in enumerate(det, 1):
        cell = ws2.cell(row=1, column=j, value=c)
        cell.font = HEAD; cell.fill = FILL; cell.border = BORD
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, c in enumerate(det, 1):
            cell = ws2.cell(row=i, column=j, value=row[c])
            cell.font = BODY; cell.border = BORD
            if c not in ("sku", "clase", "grupo"):
                cell.number_format = "#,##0"
    for j in range(1, len(det) + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 18
    ws2.freeze_panes = "A2"

    # ---- hoja 3: parámetros / trazabilidad ----
    ws3 = wb.create_sheet("Parámetros")
    ws3["A1"] = "Supuestos del modelo"; ws3["A1"].font = Font(name="Arial", bold=True, size=12, color=NAVY)
    filas = [
        ("Lead time (meses)", cfg.lead_time, "Orden → liberación SEREMI"),
        ("Horizonte (meses)", cfg.horizon, "Planificación"),
        ("Sesgo del forecast corregido", f"+{cfg.forecast_bias:.1%}", "Sobre-pronóstico al lag 6 (sección 8)"),
        ("Estiba: patrón chico", f"{cfg.stow_23} pallets exactos", "Balance de carga (proveedor)"),
        ("Estiba: patrón completo", f"{cfg.stow_full_min}-{cfg.stow_full_max} pallets", "Balance de carga (proveedor)"),
        ("Nivel de servicio A / B / C", "98% / 95% / 90%", "Z = 2,054 / 1,645 / 1,282"),
        ("Stock de seguridad", "SS = Z · σ · √L", "σ = RMSE del error de pronóstico"),
    ]
    ws3.append([]); 
    for j, h in enumerate(["Parámetro", "Valor", "Fuente / nota"], 1):
        cell = ws3.cell(row=3, column=j, value=h); cell.font = HEAD; cell.fill = FILL; cell.border = BORD
    for i, (a, b, c) in enumerate(filas, start=4):
        for j, v in enumerate([a, b, c], 1):
            cell = ws3.cell(row=i, column=j, value=v); cell.font = BODY; cell.border = BORD
    # stock de seguridad por SKU
    base = len(filas) + 6
    ws3.cell(row=base - 1, column=1, value="Stock de seguridad por SKU").font = Font(
        name="Arial", bold=True, size=11, color=NAVY)
    for j, h in enumerate(["SKU", "Clase", "Z", "σ (RMSE)", "SS = Z·σ·√L"], 1):
        cell = ws3.cell(row=base, column=j, value=h); cell.font = HEAD; cell.fill = FILL; cell.border = BORD
    for i, (s, ss) in enumerate(result["safety_stock"].items(), start=base + 1):
        sk = skus[s]
        for j, v in enumerate([s, sk.abc_class, round(sk.z(), 3), round(sk.rmse), round(ss)], 1):
            cell = ws3.cell(row=i, column=j, value=v); cell.font = BODY; cell.border = BORD
    for j, w in enumerate([30, 22, 42, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ---- hoja 4: comparación de escenarios de vida útil ----
    if results_all:
        ws4 = wb.create_sheet("Comparación escenarios")
        ws4["A1"] = "DOS SUGERENCIAS DE COMPRA — según la vida útil residual al arribo"
        ws4["A1"].font = Font(name="Arial", bold=True, size=12, color=NAVY)
        ws4["A2"] = ("BASE = promedio ponderado por unidades arribadas  ·  "
                     "CONSERVADOR = percentil P25 (asume que el lote llega más viejo). "
                     "El conservador compra menos y se expone menos a vencimientos.")
        ws4["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
        heads = ["SKU", "Clase",
                 "VU base (meses)", "Pallets BASE", "Unidades BASE", "Quiebre BASE",
                 "VU consv. (meses)", "Pallets CONSV.", "Unidades CONSV.", "Quiebre CONSV.",
                 "Δ pallets (consv. − base)"]
        for j, h in enumerate(heads, 1):
            c = ws4.cell(row=4, column=j, value=h)
            c.font = HEAD; c.fill = FILL; c.border = BORD
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        rb, rc = results_all.get("BASE"), results_all.get("CONSERVADOR")
        if rb and rc:
            db, dc_ = rb["resultados"], rc["resultados"]
            for i, sku in enumerate(sorted(db["sku"].unique()), start=5):
                gb, gc = db[db["sku"] == sku], dc_[dc_["sku"] == sku]
                pb, pc = int(gb["pallets"].sum()), int(gc["pallets"].sum())
                vals = [sku, rb["skus"][sku].abc_class,
                        rb["skus"][sku].shelf_life_months, pb,
                        int(gb["unidades_pedidas"].sum()), int(gb["quiebre_u"].sum()),
                        rc["skus"][sku].shelf_life_months, pc,
                        int(gc["unidades_pedidas"].sum()), int(gc["quiebre_u"].sum()),
                        pc - pb]
                for j, v in enumerate(vals, 1):
                    c = ws4.cell(row=i, column=j, value=v)
                    c.font = BODY; c.border = BORD
                    if j not in (1, 2):
                        c.number_format = "#,##0"
        for j, w in enumerate([20, 8, 14, 12, 14, 12, 15, 14, 15, 14, 16], 1):
            ws4.column_dimensions[get_column_letter(j)].width = w
        ws4.freeze_panes = "A5"

    wb.save(path)
    return path


def shelf_life_from_imports(
    df_imports: pd.DataFrame,
    sku_col: str = "sku",
    vu_col: str = "vida_util_restante_meses",
    qty_col: str = "unidades_arribadas",
    percentile: float = 25.0,
) -> pd.DataFrame:
    """Calcula la vida útil residual al arribo desde el REGISTRO DE IMPORTACIONES.

    Devuelve, por SKU, los dos escenarios de la restricción de vida útil (Ec. 11):

      · shelf_life_w   — promedio PONDERADO por unidades arribadas.
            VU_w = Σ(VU_i · q_i) / Σ q_i
        Un embarque de 10.000 unidades pesa más que uno de 200; es el valor
        esperado real de la vida útil con que llega el inventario.

      · shelf_life_p25 — percentil conservador (por defecto P25) ponderado por
        unidades. Protege contra los embarques que llegan MÁS VIEJOS, que son
        justamente los que generan las mermas. El promedio los diluye.

    Ambos se truncan hacia abajo a meses enteros (la restricción opera en
    períodos mensuales, y redondear hacia arriba sería optimista).
    """
    def _wpercentile(vu: np.ndarray, w: np.ndarray, q: float) -> float:
        """Percentil ponderado por unidades arribadas."""
        idx = np.argsort(vu)
        vu, w = vu[idx], w[idx]
        cw = np.cumsum(w)
        if cw[-1] <= 0:
            return float(np.percentile(vu, q))
        cutoff = (q / 100.0) * cw[-1]
        return float(vu[np.searchsorted(cw, cutoff)])

    out = []
    for sku, g in df_imports.groupby(sku_col):
        vu = g[vu_col].to_numpy(dtype=float)
        w  = g[qty_col].to_numpy(dtype=float)
        w  = np.where(np.isfinite(w) & (w > 0), w, 0.0)
        ok = np.isfinite(vu)
        vu, w = vu[ok], w[ok]
        if len(vu) == 0:
            continue
        vu_w   = float(np.average(vu, weights=w)) if w.sum() > 0 else float(vu.mean())
        vu_p25 = _wpercentile(vu, w, percentile)
        # El escenario conservador debe ser SIEMPRE ≤ el base. Cuando la
        # distribución tiene cola izquierda pesada pero poco volumen (pocos
        # embarques muy viejos), el percentil puede quedar POR ENCIMA de la
        # media ponderada —que sí absorbe esa cola—. Tomar el mínimo garantiza
        # que "conservador" signifique efectivamente una vida útil más corta.
        vu_cons = min(vu_p25, vu_w)
        out.append({
            sku_col: sku,
            "shelf_life_w":   max(1, int(np.floor(vu_w))),
            "shelf_life_p25": max(1, int(np.floor(vu_cons))),
            "vu_prom_pond":   round(vu_w, 2),
            "vu_p25":         round(vu_p25, 2),
            "vu_conservador": round(vu_cons, 2),
            "vu_min":         float(vu.min()),
            "n_embarques":    int(len(vu)),
            "unidades_hist":  float(w.sum()),
        })
    return pd.DataFrame(out)


# Escenarios de vida útil que el modelo resuelve en paralelo
SCENARIOS = {
    "BASE":        ("shelf_life_w",   "Promedio ponderado por unidades arribadas"),
    "CONSERVADOR": ("shelf_life_p25", "Percentil P25 (protege contra lotes que llegan viejos)"),
}


def solve_scenarios(
    skus: List[SKU],
    forecast: Dict[str, np.ndarray],
    expirations: Dict[str, np.ndarray],
    arrivals: Optional[Dict[str, np.ndarray]] = None,
    cfg: ModelConfig = ModelConfig(),
) -> Dict[str, Dict]:
    """Resuelve el MILP UNA VEZ POR ESCENARIO de vida útil y devuelve ambos planes.

    Entrega dos sugerencias de compra:
      · BASE        — vida útil = promedio ponderado. Menos conservador: compra más.
      · CONSERVADOR — vida útil = P25. Asume que el lote llega más viejo: compra menos,
                      pero se expone menos a vencimientos.

    El área de Compras compara ambas y decide el pedido a cargar en la plataforma.
    """
    import copy
    results = {}
    for name, (field, _desc) in SCENARIOS.items():
        skus_esc = []
        for sk in skus:
            s2 = copy.copy(sk)
            vu = getattr(sk, field, None)
            if vu is None:                       # sin dato del registro → usa el vigente
                vu = sk.shelf_life_months
            s2.shelf_life_months = int(vu)
            skus_esc.append(s2)
        results[name] = solve_milp(skus_esc, forecast, expirations, arrivals, cfg)
        results[name]["escenario"] = name
    return results


def compare_scenarios(results: Dict[str, Dict]) -> pd.DataFrame:
    """Tabla comparativa de los dos planes de compra, por SKU."""
    filas = []
    for name, res in results.items():
        df = res["resultados"]
        for sku, g in df.groupby("sku"):
            sk = res["skus"][sku]
            filas.append({
                "sku": sku,
                "escenario": name,
                "vida_util_usada": sk.shelf_life_months,
                "pallets_total": int(g["pallets"].sum()),
                "unidades_total": int(g["unidades_pedidas"].sum()),
                "n_pedidos": int((g["pallets"] > 0).sum()),
                "quiebre_u": int(g["quiebre_u"].sum()),
                "vencimiento_u": int(g["vencimiento_u"].sum()),
            })
    comp = pd.DataFrame(filas)
    return comp.pivot(index="sku", columns="escenario").swaplevel(axis=1).sort_index(axis=1)


# =============================================================================
# 5. CARGA DE DATOS DESDE BIGQUERY (esqueleto para producción)
# =============================================================================

def load_from_bigquery(project: str, dataset: str) -> Dict:
    """Esqueleto de integración con BigQuery (sección 5.3).

    En producción, reemplaza al generador sintético del demo. Cada consulta
    mapea a una tabla de los modelos intermedios del Capítulo 6:
        - parámetros SKU + clase ABC + CAP + vida útil + costos  → Bloque 3
        - forecast D̂_{p,t}                                       → tabla de forecast
        - vencimientos V_{p,t}                                    → vencimiento_producto_mes
        - inventario inicial / arribos confirmados               → fact_inventory_projection
        - σ_p (RMSE por SKU)                                      → fcst_accuracy_*_lag
    """
    raise NotImplementedError(
        "Conectar con google.cloud.bigquery y poblar SKU/forecast/expirations/"
        "arrivals desde las tablas del Capítulo 6. Ver docstring."
    )


# =============================================================================
# 6. DEMOSTRACIÓN EJECUTABLE (datos sintéticos con parámetros reales de la tesis)
# =============================================================================

def _demo_data(cfg: ModelConfig):
    """Genera un portafolio de 3 SKU (clases A, B, C) con estacionalidad y los
    parámetros de la tesis: sesgo +23 %, σ = RMSE, contenedores 23 / 40-44, etc."""
    rng = np.random.default_rng(42)
    T = cfg.horizon
    #        SKU                 mu    clase  CAP   u/caja  VU  RMSE
    #        SKU                 mu   clase CAP  u/caja VU_pond VU_p25 RMSE  grupo
    base = {"SKU-A-ALBONDIGA": (8000, "A", 1200, 12,  8, 6, 1567, "CONGELADO"),
            "SKU-B-SALMON":    (3000, "B",  800,  6, 10, 7,  800, "CONGELADO"),
            "SKU-C-MERMELADA": (1200, "C", 1500, 24, 14, 11, 363, "CONGELADO")}
    skus, forecast, expirations, arrivals = [], {}, {}, {}
    for sid, (mu, cls, cap, upc, vu_w, vu_p25, rmse, grp) in base.items():
        # demanda real con estacionalidad + ruido
        season = 1 + 0.25 * np.sin(np.linspace(0, 2 * np.pi, T))
        real = mu * season * (1 + rng.normal(0, 0.05, T))
        # el forecast nacional SOBREESTIMA en +23 % (sesgo de la sección 8)
        fc = real * (1 + cfg.forecast_bias)
        forecast[sid] = fc
        # vencimientos proyectados del stock vigente (modelo de obsolescencia)
        expirations[sid] = np.clip(rng.normal(0.02 * mu, 0.01 * mu, T), 0, None)
        # arribos confirmados en tránsito en los primeros meses
        arr = np.zeros(T); arr[:cfg.lead_time] = mu  # ~1 mes de cobertura ya en camino
        arrivals[sid] = arr
        skus.append(SKU(
            sku_id=sid, abc_class=cls, units_per_pallet=cap, units_per_case=upc,
            shelf_life_months=vu_w, shelf_life_w=vu_w, shelf_life_p25=vu_p25,
            holding_cost=120_000 / cap,             # $120.000/pallet/mes ÷ CAP
            unit_cost=3500, stockout_penalty=9000, expiry_cost=3500,
            rmse=rmse, init_inventory=1.5 * mu, container_group=grp,
        ))
    return skus, forecast, expirations, arrivals


def main():
    cfg = ModelConfig(horizon=12, lead_time=6, forecast_bias=0.232,
                      hard_safety_stock=False, solver_msg=False)
    skus, forecast, expirations, arrivals = _demo_data(cfg)

    # DOS sugerencias de compra: vida útil promedio ponderado vs. percentil P25
    todos = solve_scenarios(skus, forecast, expirations, arrivals, cfg)
    res = todos["BASE"]

    print("=" * 78)
    print(" MODELO MILP DE COMPRA — IKEA Food Chile  (demo con parámetros de tesis)")
    print("=" * 78)
    print(f" Estado del solver : {res['status']}")
    print(f" Costo total óptimo: ${res['costo_total']:,.0f} CLP")
    print(f" Sesgo corregido   : +{cfg.forecast_bias:.1%}  |  Lead time L = {cfg.lead_time} meses")
    print("\n Stock de seguridad por SKU (Ec. 4  SS = Z·σ·√L):")
    for sid, ss in res["safety_stock"].items():
        sk = res["skus"][sid]
        print(f"   {sid:18s} clase {sk.abc_class}  Z={sk.z():.3f}  σ={sk.rmse:>5.0f}  →  SS={ss:>8,.0f} u")

    df = res["resultados"]
    print("\n Plan de compra (períodos con pedido):")
    ped = df[df["pallets"] > 0][["sku", "grupo", "periodo", "pallets", "unidades_pedidas",
                                 "contenedor_grupo", "inventario_fin", "stock_seguridad"]]
    print(ped.to_string(index=False) if len(ped) else "   (sin pedidos en el horizonte)")

    tot = df.groupby("sku").agg(
        pedidos=("pallets", lambda s: (s > 0).sum()),
        pallets_tot=("pallets", "sum"),
        quiebre_tot=("quiebre_u", "sum"),
        venc_tot=("vencimiento_u", "sum"),
    )
    print("\n Resumen por SKU:")
    print(tot.to_string())

    # gráfico del modelo para el SKU clase A
    plot_sku(res, "SKU-A-ALBONDIGA", path="milp_plan_SKU-A.png")
    print("\n Figura guardada: milp_plan_SKU-A.png")

    # --- comparación de los dos escenarios de vida útil ---
    print("\n" + "=" * 78)
    print(" DOS SUGERENCIAS DE COMPRA (escenarios de vida útil residual)")
    print("=" * 78)
    for nombre, r in todos.items():
        d = r["resultados"]
        vus = {s: r["skus"][s].shelf_life_months for s in r["skus"]}
        print(f" {nombre:12s} VU={vus}  pallets={int(d['pallets'].sum()):>4}  "
              f"quiebre={int(d['quiebre_u'].sum()):>7,}  costo=${r['costo_total']:,.0f}")

    # Excel para Compras: plan BASE + hoja de comparación de escenarios
    xlsx = export_excel(res, "plan_compra_milp.xlsx", results_all=todos)
    print("\n Excel generado :", xlsx, "(incluye hoja 'Comparación escenarios')")
    return todos


if __name__ == "__main__":
    main()
